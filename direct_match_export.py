#!/usr/bin/env python3
"""
LSports ↔ TheSports 批量联赛匹配导出脚本
改进版 v2：当 TheSports 联赛的 host_country 为空时，
从联赛名称中提取国家信息，用于地区校验，消除跨国误匹配。

改进说明：
1. 新增 COUNTRY_NAME_MAP：国家名称词典，覆盖国家名、形容词、缩写、别名。
2. 新增 extract_country_from_name()：从联赛名称中提取国家信息。
3. location_veto() 升级：当 ts_country 为空时，尝试从 TS 联赛名称中提取国家再校验。
4. match_league() 升级：同国加权逻辑同时支持 ts_country 和从名称提取的国家。
"""

import pymysql
import difflib
import unicodedata
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# 数据库连接配置（使用已建立的 SSH 隧道）
# ─────────────────────────────────────────────────────────────────────────────
DB_PASSWORD = 'r74pqyYtgdjlYB41jmWA'
LS_PORT = 3309   # LSports 隧道端口
TS_PORT = 3308   # TheSports 隧道端口

def get_conn(port, db):
    return pymysql.connect(
        host='127.0.0.1', port=port,
        user='root', password=DB_PASSWORD,
        database=db, charset='utf8mb4', connect_timeout=10
    )

# ─────────────────────────────────────────────────────────────────────────────
# 已知映射表：LS tournament_id → TS competition_id
# ─────────────────────────────────────────────────────────────────────────────
KNOWN_LS_TS_MAP = {
    "football:67":    "jednm9whz0ryox8",  # Premier League (England)
    "football:8363":  "vl7oqdehlyr510j",  # LaLiga (Spain)
    "football:61":    "yl5ergphnzr8k0o",  # Ligue 1 (France)
    "football:66":    "gy0or5jhg6qwzv3",  # Bundesliga (Germany)
    "football:32644": "z8yomo4h7wq0j6l",  # UEFA Champions League
    "football:30444": "56ypq3nh0xmd7oj",  # UEFA Europa League
    "basketball:132": "49vjxm8xt4q6odg",  # NBA
}

# ─────────────────────────────────────────────────────────────────────────────
# 国家名称词典
# 格式：词/短语（小写） → 标准国家名（与 LSports ls_category 对齐）
# 覆盖：国家名、形容词形式、缩写、常见别名
# ─────────────────────────────────────────────────────────────────────────────
COUNTRY_NAME_MAP = {
    # ── 亚洲 ──
    "chinese": "China", "china": "China", "chn": "China", "cfa": "China",
    "japanese": "Japan", "japan": "Japan", "jpn": "Japan",
    "korean": "Korea", "korea": "Korea", "south korean": "South Korea",
    "south korea": "South Korea", "kor": "South Korea",
    "north korea": "North Korea", "dpr korea": "North Korea",
    "indian": "India", "india": "India", "ind": "India",
    "thai": "Thailand", "thailand": "Thailand",
    "vietnamese": "Vietnam", "vietnam": "Vietnam",
    "indonesian": "Indonesia", "indonesia": "Indonesia",
    "malaysian": "Malaysia", "malaysia": "Malaysia",
    "singaporean": "Singapore", "singapore": "Singapore",
    "myanmar": "Myanmar", "burmese": "Myanmar",
    "cambodian": "Cambodia", "cambodia": "Cambodia",
    "laos": "Laos", "laotian": "Laos",
    "philippine": "Philippines", "philippines": "Philippines",
    "filipino": "Philippines",
    "bangladeshi": "Bangladesh", "bangladesh": "Bangladesh",
    "pakistani": "Pakistan", "pakistan": "Pakistan",
    "nepali": "Nepal", "nepal": "Nepal",
    "sri lankan": "Sri Lanka", "sri lanka": "Sri Lanka",
    "iranian": "Iran", "iran": "Iran",
    "iraqi": "Iraq", "iraq": "Iraq",
    "saudi": "Saudi Arabia", "saudi arabia": "Saudi Arabia",
    "kuwaiti": "Kuwait", "kuwait": "Kuwait",
    "qatari": "Qatar", "qatar": "Qatar",
    "emirati": "UAE", "uae": "UAE", "united arab emirates": "UAE",
    "bahraini": "Bahrain", "bahrain": "Bahrain",
    "omani": "Oman", "oman": "Oman",
    "yemeni": "Yemen", "yemen": "Yemen",
    "jordanian": "Jordan", "jordan": "Jordan",
    "lebanese": "Lebanon", "lebanon": "Lebanon",
    "syrian": "Syria", "syria": "Syria",
    "uzbek": "Uzbekistan", "uzbekistan": "Uzbekistan",
    "kazakh": "Kazakhstan", "kazakhstan": "Kazakhstan",
    "tajik": "Tajikistan", "tajikistan": "Tajikistan",
    "kyrgyz": "Kyrgyzstan", "kyrgyzstan": "Kyrgyzstan",
    "turkmen": "Turkmenistan", "turkmenistan": "Turkmenistan",
    "afghan": "Afghanistan", "afghanistan": "Afghanistan",
    "mongolian": "Mongolia", "mongolia": "Mongolia",
    "hong kong": "Hong Kong",
    "macau": "Macau", "macao": "Macau",
    "taiwanese": "Taiwan", "taiwan": "Taiwan",

    # ── 欧洲 ──
    "english": "England", "england": "England",
    "scottish": "Scotland", "scotland": "Scotland",
    "welsh": "Wales", "wales": "Wales",
    "northern irish": "Northern Ireland", "northern ireland": "Northern Ireland",
    "german": "Germany", "germany": "Germany",
    "french": "France", "france": "France",
    "spanish": "Spain", "spain": "Spain",
    "italian": "Italy", "italy": "Italy",
    "portuguese": "Portugal", "portugal": "Portugal",
    "dutch": "Netherlands", "netherlands": "Netherlands", "holland": "Netherlands",
    "belgian": "Belgium", "belgium": "Belgium",
    "swiss": "Switzerland", "switzerland": "Switzerland",
    "austrian": "Austria", "austria": "Austria",
    "swedish": "Sweden", "sweden": "Sweden",
    "norwegian": "Norway", "norway": "Norway",
    "danish": "Denmark", "denmark": "Denmark",
    "finnish": "Finland", "finland": "Finland",
    "icelandic": "Iceland", "iceland": "Iceland",
    "russian": "Russia", "russia": "Russia",
    "ukrainian": "Ukraine", "ukraine": "Ukraine",
    "polish": "Poland", "poland": "Poland",
    "czech": "Czech Republic", "czechia": "Czech Republic",
    "slovak": "Slovakia", "slovakia": "Slovakia",
    "hungarian": "Hungary", "hungary": "Hungary",
    "romanian": "Romania", "romania": "Romania",
    "bulgarian": "Bulgaria", "bulgaria": "Bulgaria",
    "greek": "Greece", "greece": "Greece",
    "turkish": "Turkey", "turkey": "Turkey",
    "croatian": "Croatia", "croatia": "Croatia",
    "serbian": "Serbia", "serbia": "Serbia",
    "slovenian": "Slovenia", "slovenia": "Slovenia",
    "bosnian": "Bosnia", "bosnia": "Bosnia",
    "macedonian": "North Macedonia", "north macedonia": "North Macedonia",
    "albanian": "Albania", "albania": "Albania",
    "montenegrin": "Montenegro", "montenegro": "Montenegro",
    "kosovan": "Kosovo", "kosovo": "Kosovo",
    "belarusian": "Belarus", "belarus": "Belarus",
    "moldovan": "Moldova", "moldova": "Moldova",
    "lithuanian": "Lithuania", "lithuania": "Lithuania",
    "latvian": "Latvia", "latvia": "Latvia",
    "estonian": "Estonia", "estonia": "Estonia",
    "cypriot": "Cyprus", "cyprus": "Cyprus",
    "maltese": "Malta", "malta": "Malta",
    "luxembourgish": "Luxembourg", "luxembourg": "Luxembourg",
    "irish": "Ireland", "ireland": "Ireland",
    "faroese": "Faroe Islands", "faroe islands": "Faroe Islands",
    "azerbaijani": "Azerbaijan", "azerbaijan": "Azerbaijan",
    "armenian": "Armenia", "armenia": "Armenia",
    "georgian": "Georgia",  # 注意：Georgia 也是美国州名，但在足球语境中通常是国家
    "israeli": "Israel", "israel": "Israel",
    "andorran": "Andorra", "andorra": "Andorra",
    "liechtenstein": "Liechtenstein",
    "san marino": "San Marino",
    "gibraltar": "Gibraltar",

    # ── 美洲 ──
    "american": "USA", "usa": "USA", "us": "USA", "united states": "USA",
    "canadian": "Canada", "canada": "Canada",
    "mexican": "Mexico", "mexico": "Mexico",
    "brazilian": "Brazil", "brazil": "Brazil", "bra": "Brazil",
    "argentine": "Argentina", "argentinian": "Argentina", "argentina": "Argentina",
    "colombian": "Colombia", "colombia": "Colombia",
    "chilean": "Chile", "chile": "Chile",
    "peruvian": "Peru", "peru": "Peru",
    "venezuelan": "Venezuela", "venezuela": "Venezuela",
    "ecuadorian": "Ecuador", "ecuador": "Ecuador",
    "bolivian": "Bolivia", "bolivia": "Bolivia",
    "paraguayan": "Paraguay", "paraguay": "Paraguay",
    "uruguayan": "Uruguay", "uruguay": "Uruguay",
    "costa rican": "Costa Rica", "costa rica": "Costa Rica",
    "guatemalan": "Guatemala", "guatemala": "Guatemala",
    "honduran": "Honduras", "honduras": "Honduras",
    "salvadoran": "El Salvador", "el salvador": "El Salvador",
    "nicaraguan": "Nicaragua", "nicaragua": "Nicaragua",
    "panamanian": "Panama", "panama": "Panama",
    "cuban": "Cuba", "cuba": "Cuba",
    "jamaican": "Jamaica", "jamaica": "Jamaica",
    "trinidadian": "Trinidad and Tobago", "trinidad": "Trinidad and Tobago",

    # ── 非洲 ──
    "south african": "South Africa", "south africa": "South Africa",
    "nigerian": "Nigeria", "nigeria": "Nigeria",
    "ghanaian": "Ghana", "ghana": "Ghana", "gha": "Ghana",
    "kenyan": "Kenya", "kenya": "Kenya",
    "tanzanian": "Tanzania", "tanzania": "Tanzania",
    "ugandan": "Uganda", "uganda": "Uganda",
    "rwandan": "Rwanda", "rwanda": "Rwanda",
    "ethiopian": "Ethiopia", "ethiopia": "Ethiopia",
    "moroccan": "Morocco", "morocco": "Morocco",
    "algerian": "Algeria", "algeria": "Algeria", "alg": "Algeria",
    "tunisian": "Tunisia", "tunisia": "Tunisia",
    "egyptian": "Egypt", "egypt": "Egypt",
    "libyan": "Libya", "libya": "Libya",
    "senegalese": "Senegal", "senegal": "Senegal",
    "ivorian": "Ivory Coast", "ivory coast": "Ivory Coast",
    "cameroonian": "Cameroon", "cameroon": "Cameroon",
    "congolese": "Congo", "congo": "Congo",
    "zambian": "Zambia", "zambia": "Zambia",
    "zimbabwean": "Zimbabwe", "zimbabwe": "Zimbabwe",
    "mozambican": "Mozambique", "mozambique": "Mozambique",
    "angolan": "Angola", "angola": "Angola",
    "sudanese": "Sudan", "sudan": "Sudan",
    "malian": "Mali", "mali": "Mali",
    "burkinabe": "Burkina Faso", "burkina faso": "Burkina Faso",
    "beninese": "Benin", "benin": "Benin",
    "togolese": "Togo", "togo": "Togo",
    "liberian": "Liberia", "liberia": "Liberia",
    "sierra leonean": "Sierra Leone", "sierra leone": "Sierra Leone",
    "namibian": "Namibia", "namibia": "Namibia",
    "botswanan": "Botswana", "botswana": "Botswana",
    "lesotho": "Lesotho",
    "swazi": "Eswatini", "eswatini": "Eswatini",
    "malawian": "Malawi", "malawi": "Malawi",

    # ── 大洋洲 ──
    "australian": "Australia", "australia": "Australia", "aus": "Australia",
    "new zealand": "New Zealand", "new zealander": "New Zealand",
    "fijian": "Fiji", "fiji": "Fiji",
    "papua new guinea": "Papua New Guinea",
    "solomon islands": "Solomon Islands",
    "vanuatu": "Vanuatu",
    "tahitian": "Tahiti", "tahiti": "Tahiti",

    # ── 补充：缩写 & 遗漏国家 ──
    # 缩写
    "rus": "Russia", "fra": "France", "ita": "Italy", "spa": "Spain",
    "por": "Portugal", "arg": "Argentina", "mex": "Mexico",
    "par": "Paraguay", "rom": "Romania", "svk": "Slovakia",
    "ire": "Ireland", "bos": "Bosnia", "tha": "Thailand",
    "alg": "Algeria",
    # 遗漏国家
    "niger": "Niger",  # 注意：不同于 Nigeria
    "brunei": "Brunei",
    "bhutan": "Bhutan",
    "mauritania": "Mauritania",
    "aruba": "Aruba",
    "grenada": "Grenada",
    "gambia": "Gambia",
    "barbados": "Barbados",
    "burundi": "Burundi",
    "malawi": "Malawi",
    "djibouti": "Djibouti",
    "eritrea": "Eritrea",
    "somalia": "Somalia",
    "comoros": "Comoros",
    "seychelles": "Seychelles",
    "maldives": "Maldives",
    "timor leste": "Timor-Leste", "timor": "Timor-Leste",
    "macanese": "Macau",
    "costarica": "Costa Rica",  # 无空格写法
    "georgia": "Georgia",
    "liechtenstein": "Liechtenstein",
    "andorra": "Andorra",
    "faroe": "Faroe Islands",
    "kosovo": "Kosovo",
    "puerto rico": "Puerto Rico",
    "guadeloupe": "Guadeloupe",
    "martinique": "Martinique",
    "reunion": "Reunion",
    "curacao": "Curacao",
    "bermuda": "Bermuda",
    "cayman islands": "Cayman Islands",
    "haiti": "Haiti",
    "dominican": "Dominican Republic", "dominican republic": "Dominican Republic",
    "belize": "Belize",
    "suriname": "Suriname",
    "guyana": "Guyana",
    "french guiana": "French Guiana",
    "saint lucia": "Saint Lucia",
    "saint kitts": "Saint Kitts and Nevis",
    "antigua": "Antigua and Barbuda",
    "dominica": "Dominica",
    "virgin islands": "Virgin Islands",
    "montserrat": "Montserrat",
    "anguilla": "Anguilla",
    "turks": "Turks and Caicos",
}

# ─────────────────────────────────────────────────────────────────────────────
# 名称归一化与相似度
# ─────────────────────────────────────────────────────────────────────────────

def normalize_name(s: str) -> str:
    """归一化名称：去变音符号、转小写、去标点、合并空格"""
    if not s:
        return ''
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    s = s.lower()
    s = re.sub(r'[.\-_,\'\"·]', ' ', s)
    s = ' '.join(s.split())
    return s


def seq_similarity(a: str, b: str) -> float:
    na, nb = normalize_name(a), normalize_name(b)
    if not na and not nb:
        return 1.0
    if not na or not nb:
        return 0.0
    return difflib.SequenceMatcher(None, na, nb).ratio()


def jaccard_similarity(a: str, b: str) -> float:
    na, nb = normalize_name(a), normalize_name(b)
    set_a = set(na.split())
    set_b = set(nb.split())
    if not set_a and not set_b:
        return 1.0
    if not set_a or not set_b:
        return 0.0
    return len(set_a & set_b) / len(set_a | set_b)


def name_similarity(a: str, b: str) -> float:
    return max(jaccard_similarity(a, b), seq_similarity(a, b))


# ─────────────────────────────────────────────────────────────────────────────
# 国家/地区校验
# ─────────────────────────────────────────────────────────────────────────────

# 洲际/国际赛事关键词
INTERNATIONAL_KEYWORDS = {
    'world', 'international', 'europe', 'europa', 'asia', 'africa',
    'america', 'oceania', 'concacaf', 'conmebol', 'afc', 'caf',
    'uefa', 'fifa', 'ofc', 'waff', 'saff', 'eaff', 'cosafa',
    'south america', 'north america', 'central america',
}


def is_international_category(name: str) -> bool:
    """判断地区名称是否属于洲际/国际赛事"""
    if not name:
        return False
    norm = normalize_name(name)
    if norm in INTERNATIONAL_KEYWORDS:
        return True
    tokens = set(norm.split())
    return bool(tokens & INTERNATIONAL_KEYWORDS)


def extract_country_from_name(league_name: str) -> str:
    """
    从联赛名称中提取国家信息。
    策略：
    1. 优先匹配多词短语（如 "South Africa", "New Zealand"）
    2. 再匹配单词（国家名、形容词、缩写）
    3. 仅匹配名称开头的词（前3个词），避免误提取（如 "Copa America" 不提取 America）

    返回标准国家名，未找到则返回空字符串。
    """
    if not league_name:
        return ''

    norm = normalize_name(league_name)
    tokens = norm.split()

    if not tokens:
        return ''

    # 先尝试匹配前3个词组成的短语（最长优先）
    for length in (3, 2, 1):
        if len(tokens) < length:
            continue
        phrase = ' '.join(tokens[:length])
        if phrase in COUNTRY_NAME_MAP:
            return COUNTRY_NAME_MAP[phrase]

    return ''


# 预计算：为每个 TS 联赛提取国家（避免重复计算）
_ts_extracted_country_cache: dict = {}


def get_effective_ts_country(comp: dict) -> str:
    """
    获取 TS 联赛的有效国家信息：
    优先使用 host_country，若为空则从联赛名称中提取。
    """
    ts_id = comp.get('competition_id', '')
    if ts_id in _ts_extracted_country_cache:
        return _ts_extracted_country_cache[ts_id]

    country = comp.get('host_country', '') or ''
    if not country:
        country = extract_country_from_name(comp.get('name', ''))

    _ts_extracted_country_cache[ts_id] = country
    return country


def location_veto(ls_category: str, effective_ts_country: str) -> bool:
    """
    判断是否应否决该联赛匹配（跨国误匹配检测）。
    effective_ts_country 已经是经过名称提取后的有效国家信息。
    返回 True 表示地区明显不匹配，应跳过该候选。
    """
    if not ls_category or not effective_ts_country:
        return False
    if is_international_category(ls_category) or is_international_category(effective_ts_country):
        return False
    loc_sim = name_similarity(ls_category, effective_ts_country)
    return loc_sim < 0.4


# ─────────────────────────────────────────────────────────────────────────────
# 联赛匹配核心函数
# ─────────────────────────────────────────────────────────────────────────────

def match_league(ls_name: str, ls_category: str, ts_competitions: list,
                 sport: str = 'football', ls_id: str = '') -> dict:
    """
    为一个 LSports 联赛在 TheSports 中找到最佳匹配。

    改进 v2：
    - 使用 get_effective_ts_country() 获取 TS 联赛的有效国家（含从名称提取）
    - location_veto() 基于有效国家进行强约束校验
    - 同国加权逻辑同时支持 host_country 和从名称提取的国家

    返回 dict: {ts_id, ts_name, ts_country, ts_extracted_country, score, rule, matched}
    """
    result = {
        'ts_id': '', 'ts_name': '', 'ts_country': '', 'ts_extracted_country': '',
        'score': 0.0, 'rule': 'NO_MATCH', 'matched': False
    }

    # 1. 已知映射
    map_key = f"{sport}:{ls_id}"
    if ls_id and map_key in KNOWN_LS_TS_MAP:
        ts_id = KNOWN_LS_TS_MAP[map_key]
        for comp in ts_competitions:
            if comp['competition_id'] == ts_id:
                eff_country = get_effective_ts_country(comp)
                result.update({
                    'ts_id': comp['competition_id'],
                    'ts_name': comp['name'],
                    'ts_country': comp.get('host_country', ''),
                    'ts_extracted_country': eff_country,
                    'score': 1.0, 'rule': 'KNOWN', 'matched': True
                })
                return result
        result.update({'ts_id': ts_id, 'score': 1.0, 'rule': 'KNOWN', 'matched': True})
        return result

    # 2. 名称相似度匹配（含国家/地区强约束）
    best_score = 0.0
    best_comp = None
    best_eff_country = ''

    for comp in ts_competitions:
        # 获取有效国家（host_country 或从名称提取）
        eff_country = get_effective_ts_country(comp)

        # 强约束：地区明显不匹配时跳过
        if location_veto(ls_category, eff_country):
            continue

        # 计算联赛名称相似度
        base_score = name_similarity(ls_name, comp['name'])

        # 同国加权 / 跨国惩罚
        if ls_category and eff_country:
            loc_sim = name_similarity(ls_category, eff_country)
            if loc_sim >= 0.6:
                # 同国：地区相似度作为加权项提升置信度
                base_score = base_score * 0.75 + 0.25 * loc_sim
            elif 0.4 <= loc_sim < 0.6:
                # 模糊地区（如 Ireland/Iraq、Japan/Jordan）：施加惩罚
                # 惩罚力度与地区相似度成反比：loc_sim 越低，惩罚越重
                penalty = 0.70 + 0.30 * (loc_sim - 0.4) / 0.2  # [0.4,0.6) → [0.70, 1.00)
                base_score = base_score * penalty
            # loc_sim < 0.4 已被 location_veto() 拦截，不会到达此处
        elif ls_category and not eff_country and not is_international_category(ls_category):
            # LS 有地区信息但 TS 完全无国家信息（名称中也无法提取）
            # 对低置信度区间施加轻度惩罚
            if base_score < 0.75:
                base_score = base_score * 0.90

        if base_score > best_score:
            best_score = base_score
            best_comp = comp
            best_eff_country = eff_country

    if best_comp is None:
        return result

    # 根据分数确定匹配规则
    if best_score >= 0.85:
        rule = 'NAME_HI'
    elif best_score >= 0.70:
        rule = 'NAME_MED'
    elif best_score >= 0.55:
        rule = 'NAME_LOW'
    else:
        return result

    result.update({
        'ts_id': best_comp['competition_id'],
        'ts_name': best_comp['name'],
        'ts_country': best_comp.get('host_country', ''),
        'ts_extracted_country': best_eff_country,
        'score': round(best_score, 4),
        'rule': rule,
        'matched': True
    })
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 数据加载
# ─────────────────────────────────────────────────────────────────────────────

def load_ls_tournaments(sport_id: int = 6046) -> list:
    """加载 LSports 联赛列表（含地区信息）"""
    conn = get_conn(LS_PORT, 'test-xp-lsports')
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        cur.execute("""
            SELECT t.tournament_id, t.name as tournament_name,
                   COALESCE(c.name, '') as location
            FROM ls_tournament_en t
            LEFT JOIN ls_category_en c
                ON t.category_id = c.category_id
                AND c.sport_id = %s
            WHERE t.sport_id = %s
            GROUP BY t.tournament_id, t.name, c.name
            ORDER BY t.tournament_id
        """, (str(sport_id), int(sport_id)))
        rows = cur.fetchall()
    conn.close()
    return rows


def load_ts_competitions(sport: str = 'football') -> list:
    """加载 TheSports 联赛列表"""
    conn = get_conn(TS_PORT, 'test-thesports-db')
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        if sport == 'football':
            cur.execute("""
                SELECT competition_id, name,
                       COALESCE(host_country, '') as host_country
                FROM ts_fb_competition
            """)
        elif sport == 'basketball':
            cur.execute("""
                SELECT competition_id, name, '' as host_country
                FROM ts_bb_competition
            """)
        rows = cur.fetchall()
    conn.close()
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# 主流程：批量匹配并导出 Excel
# ─────────────────────────────────────────────────────────────────────────────

def run_batch_match(sport: str = 'football', output_path: str = None):
    """执行批量联赛匹配并导出结果"""
    # 清空缓存（多次调用时重置）
    _ts_extracted_country_cache.clear()

    print(f"[{datetime.now().strftime('%H:%M:%S')}] 加载 LSports 联赛...")
    sport_id_map = {'football': 6046, 'basketball': 48242}
    ls_tours = load_ls_tournaments(sport_id_map.get(sport, 6046))
    print(f"  LSports 联赛数: {len(ls_tours)}")

    print(f"[{datetime.now().strftime('%H:%M:%S')}] 加载 TheSports 联赛...")
    ts_comps = load_ts_competitions(sport)
    print(f"  TheSports 联赛数: {len(ts_comps)}")

    # 预计算所有 TS 联赛的有效国家（避免重复提取）
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 预计算 TS 联赛国家信息...")
    extracted_count = 0
    for comp in ts_comps:
        eff = get_effective_ts_country(comp)
        if eff and not (comp.get('host_country', '') or ''):
            extracted_count += 1
    print(f"  从名称中提取到国家信息: {extracted_count}/{len(ts_comps)} 条")

    print(f"[{datetime.now().strftime('%H:%M:%S')}] 开始批量匹配...")
    results = []
    matched_count = 0
    known_count = 0

    for i, tour in enumerate(ls_tours):
        ls_id = str(tour['tournament_id'])
        ls_name = tour['tournament_name'] or ''
        ls_category = tour['location'] or ''

        match = match_league(ls_name, ls_category, ts_comps, sport, ls_id)

        if match['matched']:
            matched_count += 1
            if match['rule'] == 'KNOWN':
                known_count += 1

        results.append({
            'ls_tournament_id': ls_id,
            'ls_name': ls_name,
            'ls_category': ls_category,
            'ts_competition_id': match['ts_id'],
            'ts_name': match['ts_name'],
            'ts_country': match['ts_country'],
            'ts_extracted_country': match.get('ts_extracted_country', ''),
            'score': match['score'],
            'rule': match['rule'],
            'matched': match['matched'],
        })

        if (i + 1) % 100 == 0:
            print(f"  进度: {i+1}/{len(ls_tours)}, 已匹配: {matched_count}")

    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 匹配完成:")
    print(f"  总联赛数: {len(ls_tours)}")
    print(f"  已匹配:   {matched_count} ({matched_count/len(ls_tours)*100:.1f}%)")
    print(f"  已知映射: {known_count}")
    print(f"  未匹配:   {len(ls_tours) - matched_count}")

    if output_path is None:
        output_path = f'/home/ubuntu/lsports_ts_match_result_{sport}.xlsx'

    export_excel(results, output_path, sport)
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 结果已导出: {output_path}")
    return results, output_path


def export_excel(results: list, output_path: str, sport: str):
    """将匹配结果导出为 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = f'LS-TS Match ({sport})'

    headers = [
        'LS Tournament ID', 'LS League Name', 'LS Category (Location)',
        'TS Competition ID', 'TS League Name', 'TS Country (DB)',
        'TS Country (Extracted)', 'Score', 'Match Rule', 'Matched'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    fill_known = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_hi    = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')
    fill_med   = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_low   = PatternFill(start_color='FFD7B5', end_color='FFD7B5', fill_type='solid')
    fill_none  = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    fill_map = {
        'KNOWN': fill_known, 'NAME_HI': fill_hi,
        'NAME_MED': fill_med, 'NAME_LOW': fill_low, 'NO_MATCH': fill_none,
    }

    for row in results:
        ws.append([
            row['ls_tournament_id'],
            row['ls_name'],
            row['ls_category'],
            row['ts_competition_id'],
            row['ts_name'],
            row['ts_country'],
            row.get('ts_extracted_country', ''),
            row['score'],
            row['rule'],
            'YES' if row['matched'] else 'NO',
        ])
        fill = fill_map.get(row['rule'], fill_none)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    col_widths = [18, 45, 25, 22, 45, 20, 20, 8, 12, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    # 统计 Sheet
    ws2 = wb.create_sheet('统计')
    matched = [r for r in results if r['matched']]
    rule_counts = {}
    for r in results:
        rule_counts[r['rule']] = rule_counts.get(r['rule'], 0) + 1

    ws2.append(['指标', '数值'])
    ws2.append(['总联赛数', len(results)])
    ws2.append(['已匹配', len(matched)])
    ws2.append(['匹配率', f"{len(matched)/len(results)*100:.1f}%" if results else '0%'])
    ws2.append([''])
    ws2.append(['匹配规则', '数量'])
    for rule, cnt in sorted(rule_counts.items()):
        ws2.append([rule, cnt])

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys
    sport = sys.argv[1] if len(sys.argv) > 1 else 'football'
    output = sys.argv[2] if len(sys.argv) > 2 else None
    run_batch_match(sport=sport, output_path=output)
