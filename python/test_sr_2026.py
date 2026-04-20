"""
test_sr_2026.py — SR↔TS 2026 年算法测试脚本（最新算法，离线版 v2）
================================================================
复现 UniversalEngine（P0~P3 全量优化）的核心匹配逻辑：
  - 高斯时间衰减连续模糊时间窗口（σ=3600s/10800s/43200s）
  - 主客场颠倒否决
  - 联赛级别名学习（TeamAliasIndex）
  - L1~L5 + L4b + L6 七级匹配策略
  - 已知映射（KnownLeagueMap）优先

数据策略：
  - 优先使用 ts_events_2026.json 中有数据的联赛
  - 对 ts_events_2026.json 中无数据但 GT 中有记录的联赛，
    从 sr_ts_ground_truth.json 重建 TS 候选集（包含干扰项）进行评估
  - 这样可以对全部 14 个 GT 联赛跑完整评估

用法：
  python3 python/test_sr_2026.py
  python3 python/test_sr_2026.py --tier hot
  python3 python/test_sr_2026.py --league sr:tournament:17
  python3 python/test_sr_2026.py --output python/data/sr_2026_test_results.xlsx
"""

import json
import math
import re
import unicodedata
import difflib
import argparse
import os
import bisect
from collections import defaultdict, Counter
from datetime import datetime, timezone

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')

# ─────────────────────────────────────────────────────────────────────────────
# 已知联赛映射（与 internal/matcher/league.go KnownLeagueMap 保持同步）
# ─────────────────────────────────────────────────────────────────────────────
KNOWN_LEAGUE_MAP = {
    # 足球热门
    "football:sr:tournament:17":  "jednm9whz0ryox8",
    "football:sr:tournament:8":   "vl7oqdehlyr510j",
    "football:sr:tournament:35":  "gy0or5jhg6qwzv3",
    "football:sr:tournament:23":  "4zp5rzghp5q82w1",
    "football:sr:tournament:34":  "yl5ergphnzr8k0o",
    "football:sr:tournament:7":   "z8yomo4h7wq0j6l",
    "football:sr:tournament:679": "56ypq3nh0xmd7oj",
    # 足球常规
    "football:sr:tournament:18":  "l965mkyh32r1ge4",
    "football:sr:tournament:37":  "vl7oqdeheyr510j",
    "football:sr:tournament:238": "gx7lm7phpnm2wdk",
    "football:sr:tournament:52":  "8y39mp1h6jmojxg",
    "football:sr:tournament:203": "8y39mp1hwxmojxg",
    "football:sr:tournament:11":  "9vjxm8gh22r6odg",
    "football:sr:tournament:242": "kn54qllhg2qvy9d",
    "football:sr:tournament:325": "4zp5rzgh9zq82w1",
    "football:sr:tournament:955": "z318q66hl1qo9jd",
    # 篮球热门
    "basketball:sr:tournament:132": "49vjxm8xt4q6odg",
    "basketball:sr:tournament:138": "jednm9ktd5ryox8",
    # 篮球常规
    "basketball:sr:tournament:390": "kjw2r02t6xqz84o",
    "basketball:sr:tournament:176": "v2y8m4ptx1ml074",
    "basketball:sr:tournament:131": "v2y8m4ptdeml074",
    "basketball:sr:tournament:53":  "x4zp5rzkt1r82w1",
    "basketball:sr:tournament:54":  "0l965mk8tom1ge4",
}

# SR 2026 联赛配置（热门+常规，与 cmd/server/main.go sr2026Leagues 保持同步）
SR_2026_LEAGUES = [
    # 足球热门
    ("sr:tournament:17",  "football",   "hot"),
    ("sr:tournament:8",   "football",   "hot"),
    ("sr:tournament:35",  "football",   "hot"),
    ("sr:tournament:23",  "football",   "hot"),
    ("sr:tournament:34",  "football",   "hot"),
    ("sr:tournament:7",   "football",   "hot"),
    ("sr:tournament:679", "football",   "hot"),
    # 足球常规
    ("sr:tournament:18",  "football",   "regular"),
    ("sr:tournament:242", "football",   "regular"),
    ("sr:tournament:203", "football",   "regular"),
    ("sr:tournament:325", "football",   "regular"),
    ("sr:tournament:37",  "football",   "regular"),
    ("sr:tournament:52",  "football",   "regular"),
    ("sr:tournament:238", "football",   "regular"),
    ("sr:tournament:11",  "football",   "regular"),
    ("sr:tournament:955", "football",   "regular"),
    # 篮球热门
    ("sr:tournament:132", "basketball", "hot"),
    ("sr:tournament:138", "basketball", "hot"),
    # 篮球常规
    ("sr:tournament:176", "basketball", "regular"),
    ("sr:tournament:131", "basketball", "regular"),
    ("sr:tournament:53",  "basketball", "regular"),
    ("sr:tournament:54",  "basketball", "regular"),
    ("sr:tournament:390", "basketball", "regular"),
]

# ─────────────────────────────────────────────────────────────────────────────
# 名称归一化与相似度（复现 internal/matcher/name.go）
# ─────────────────────────────────────────────────────────────────────────────

_norm_cache = {}

def normalize_name(s: str) -> str:
    if s in _norm_cache:
        return _norm_cache[s]
    if not s:
        _norm_cache[s] = ''
        return ''
    r = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    r = r.lower()
    r = re.sub(r"[.\-_,\'\"·|()]", ' ', r)
    r = re.sub(r'\bfc\b|\bsc\b|\bac\b|\bfk\b|\bsk\b', '', r)
    r = ' '.join(r.split())
    _norm_cache[s] = r
    return r

def jaro_winkler(s1: str, s2: str) -> float:
    if s1 == s2:
        return 1.0
    l1, l2 = len(s1), len(s2)
    if l1 == 0 or l2 == 0:
        return 0.0
    match_dist = max(l1, l2) // 2 - 1
    if match_dist < 0:
        match_dist = 0
    s1_matches = [False] * l1
    s2_matches = [False] * l2
    matches = 0
    for i in range(l1):
        start = max(0, i - match_dist)
        end = min(i + match_dist + 1, l2)
        for j in range(start, end):
            if s2_matches[j] or s1[i] != s2[j]:
                continue
            s1_matches[i] = True
            s2_matches[j] = True
            matches += 1
            break
    if matches == 0:
        return 0.0
    k = 0
    transpositions = 0
    for i in range(l1):
        if not s1_matches[i]:
            continue
        while not s2_matches[k]:
            k += 1
        if s1[i] != s2[k]:
            transpositions += 1
        k += 1
    jaro = (matches / l1 + matches / l2 + (matches - transpositions / 2) / matches) / 3
    prefix = 0
    for i in range(min(4, l1, l2)):
        if s1[i] == s2[i]:
            prefix += 1
        else:
            break
    return jaro + prefix * 0.1 * (1 - jaro)

def team_name_similarity(a: str, b: str) -> float:
    na, nb = normalize_name(a), normalize_name(b)
    if not na and not nb:
        return 1.0
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    sa, sb = set(na.split()), set(nb.split())
    j = len(sa & sb) / len(sa | sb) if (sa | sb) else 0.0
    seq = difflib.SequenceMatcher(None, na, nb).ratio()
    jw = jaro_winkler(na, nb)
    return max(j, seq, jw)

# ─────────────────────────────────────────────────────────────────────────────
# 高斯时间衰减（复现 gaussianTimeFactor）
# ─────────────────────────────────────────────────────────────────────────────

def gaussian_time_factor(time_diff_sec: float, sigma: float) -> float:
    return math.exp(-(time_diff_sec ** 2) / (2 * sigma ** 2))

# 匹配策略配置（复现 levelConfigs）
LEVEL_CONFIGS = [
    # (name, sigma, time_weight, name_weight, name_thresh, score_thresh, max_time_sec, require_alias)
    ("L1", 3600,   0.30, 0.70, 0.40, 0.50, 10800,   False),
    ("L2", 10800,  0.15, 0.85, 0.65, 0.60, 32400,   False),
    ("L3", 43200,  0.00, 1.00, 0.75, 0.70, 86400,   False),
    ("L4", 43200,  0.00, 1.00, 0.85, 0.80, 259200,  True),
    ("L5", 999999, 0.00, 1.00, 0.90, 0.90, 2592000, False),
]

# ─────────────────────────────────────────────────────────────────────────────
# 占位符队名检测
# ─────────────────────────────────────────────────────────────────────────────
PLACEHOLDER_PATTERNS = [
    re.compile(r'^WQF', re.IGNORECASE),
    re.compile(r'^Winner R\d+', re.IGNORECASE),
    re.compile(r'^Winner of', re.IGNORECASE),
    re.compile(r'^TBD$', re.IGNORECASE),
    re.compile(r'^TBA$', re.IGNORECASE),
]

def is_placeholder(name: str) -> bool:
    return any(p.search(name) for p in PLACEHOLDER_PATTERNS)

# ─────────────────────────────────────────────────────────────────────────────
# TeamAliasIndex（复现 internal/matcher/event.go TeamAliasIndex）
# ─────────────────────────────────────────────────────────────────────────────

class TeamAliasIndex:
    LEARN_THRESHOLD = 0.50
    APPLY_SCORE = 0.92

    def __init__(self):
        self.votes = defaultdict(lambda: defaultdict(int))
        self.alias = {}
        self.alias_rev = {}

    def learn(self, sr_home_id, sr_away_id, ts_home_id, ts_away_id,
              sr_home_name, sr_away_name, ts_home_name, ts_away_name, name_sim):
        if name_sim < self.LEARN_THRESHOLD:
            return
        fwd = team_name_similarity(sr_home_name, ts_home_name) + team_name_similarity(sr_away_name, ts_away_name)
        rev = team_name_similarity(sr_home_name, ts_away_name) + team_name_similarity(sr_away_name, ts_home_name)
        pairs = [(sr_home_id, ts_home_id), (sr_away_id, ts_away_id)] if fwd >= rev \
                else [(sr_home_id, ts_away_id), (sr_away_id, ts_home_id)]
        for sr_id, ts_id in pairs:
            if not sr_id or not ts_id:
                continue
            self.votes[sr_id][ts_id] += 1
            best_ts = max(self.votes[sr_id], key=self.votes[sr_id].get)
            if self.votes[sr_id][best_ts] >= 1:
                self.alias[sr_id] = best_ts
                self.alias_rev[best_ts] = sr_id

    def lookup(self, sr_team_id: str) -> str:
        return self.alias.get(sr_team_id, '')

# ─────────────────────────────────────────────────────────────────────────────
# 核心匹配函数
# ─────────────────────────────────────────────────────────────────────────────

def match_event_pair(sr_ev, ts_ev, alias_idx: TeamAliasIndex, team_id_map: dict) -> tuple:
    """返回 (level, confidence, time_diff_sec, name_sim)"""
    time_diff = abs(sr_ev['scheduled_unix'] - ts_ev['match_time'])
    sr_home, sr_away = sr_ev['home_name'], sr_ev['away_name']
    ts_home, ts_away = ts_ev['home_name'], ts_ev['away_name']

    # 主客场方向判断
    fwd_sim = (team_name_similarity(sr_home, ts_home) + team_name_similarity(sr_away, ts_away)) / 2
    rev_sim = (team_name_similarity(sr_home, ts_away) + team_name_similarity(sr_away, ts_home)) / 2
    if rev_sim > fwd_sim + 0.15:
        return None, 0.0, time_diff, 0.0  # 主客场颠倒，否决

    name_sim = fwd_sim

    # 别名加成
    alias_home = alias_idx.lookup(sr_ev['home_id'])
    alias_away = alias_idx.lookup(sr_ev['away_id'])
    has_alias = (alias_home == ts_ev['home_id'] and alias_home != '') or \
                (alias_away == ts_ev['away_id'] and alias_away != '')
    if has_alias:
        name_sim = max(name_sim, TeamAliasIndex.APPLY_SCORE)

    # L4b 球队ID精确匹配
    if team_id_map:
        mapped_home = team_id_map.get(sr_ev['home_id'], '')
        mapped_away = team_id_map.get(sr_ev['away_id'], '')
        if mapped_home == ts_ev['home_id'] and mapped_away == ts_ev['away_id']:
            return 'L4b', 0.95, time_diff, name_sim

    # 占位符 L6
    if is_placeholder(sr_home) or is_placeholder(sr_away):
        if time_diff <= 300:
            return 'L6', 0.70, time_diff, name_sim
        return None, 0.0, time_diff, name_sim

    # L1~L5
    for level_name, sigma, tw, nw, name_thresh, score_thresh, max_time, req_alias in LEVEL_CONFIGS:
        if time_diff > max_time:
            continue
        if req_alias and not has_alias:
            continue
        if name_sim < name_thresh:
            continue
        s_time = gaussian_time_factor(time_diff, sigma)
        score = tw * s_time + nw * name_sim
        if score >= score_thresh:
            return level_name, score, time_diff, name_sim

    return None, 0.0, time_diff, name_sim


def match_events_for_league(sr_events: list, ts_events: list) -> list:
    """对一个联赛的所有 SR 事件执行匹配，返回匹配结果列表"""
    alias_idx = TeamAliasIndex()

    # 构建 TS 时间索引
    ts_sorted = sorted(ts_events, key=lambda e: e['match_time'])
    ts_times = [e['match_time'] for e in ts_sorted]

    def find_ts_candidates(sr_unix, window_sec=259200):
        lo, hi = sr_unix - window_sec, sr_unix + window_sec
        i_lo = bisect.bisect_left(ts_times, lo)
        i_hi = bisect.bisect_right(ts_times, hi)
        return ts_sorted[i_lo:i_hi]

    results = []
    ts_used = set()

    # 第一轮匹配
    for sr_ev in sr_events:
        candidates = find_ts_candidates(sr_ev['scheduled_unix'])
        best_level, best_conf, best_ts, best_tdiff, best_nsim = None, 0.0, None, 0, 0.0
        for ts_ev in candidates:
            if ts_ev['match_id'] in ts_used:
                continue
            level, conf, tdiff, nsim = match_event_pair(sr_ev, ts_ev, alias_idx, {})
            if level and conf > best_conf:
                best_level, best_conf, best_ts, best_tdiff, best_nsim = level, conf, ts_ev, tdiff, nsim
        if best_ts:
            ts_used.add(best_ts['match_id'])
            alias_idx.learn(
                sr_ev['home_id'], sr_ev['away_id'],
                best_ts['home_id'], best_ts['away_id'],
                sr_ev['home_name'], sr_ev['away_name'],
                best_ts['home_name'], best_ts['away_name'],
                best_nsim,
            )
        results.append({
            'sr_event_id': sr_ev['event_id'],
            'sr_home': sr_ev['home_name'],
            'sr_away': sr_ev['away_name'],
            'sr_time': sr_ev['scheduled'],
            'ts_match_id': best_ts['match_id'] if best_ts else '',
            'ts_home': best_ts['home_name'] if best_ts else '',
            'ts_away': best_ts['away_name'] if best_ts else '',
            'ts_time': best_ts.get('match_time_str', '') if best_ts else '',
            'level': best_level or 'UNMATCHED',
            'confidence': best_conf,
            'time_diff_sec': best_tdiff,
            'name_sim': best_nsim,
            'matched': best_ts is not None,
        })

    # 推导球队 ID 映射（用于第二轮 L4b）
    team_id_map = {}
    vote_map = defaultdict(lambda: defaultdict(int))
    sr_ev_dict = {e['event_id']: e for e in sr_events}
    ts_ev_dict = {e['match_id']: e for e in ts_events}
    for r in results:
        if not r['matched']:
            continue
        sr_ev = sr_ev_dict.get(r['sr_event_id'])
        ts_ev = ts_ev_dict.get(r['ts_match_id'])
        if sr_ev and ts_ev:
            vote_map[sr_ev['home_id']][ts_ev['home_id']] += 1
            vote_map[sr_ev['away_id']][ts_ev['away_id']] += 1
    for sr_id, ts_votes in vote_map.items():
        best_ts_id = max(ts_votes, key=ts_votes.get)
        if ts_votes[best_ts_id] >= 2:
            team_id_map[sr_id] = best_ts_id

    # 第二轮匹配（L4b 球队ID兜底）
    if team_id_map:
        ts_used2 = {r['ts_match_id'] for r in results if r['matched']}
        for r in results:
            if r['matched']:
                continue
            sr_ev = sr_ev_dict.get(r['sr_event_id'])
            if not sr_ev:
                continue
            candidates = find_ts_candidates(sr_ev['scheduled_unix'])
            for ts_ev in candidates:
                if ts_ev['match_id'] in ts_used2:
                    continue
                level, conf, tdiff, nsim = match_event_pair(sr_ev, ts_ev, alias_idx, team_id_map)
                if level == 'L4b':
                    ts_used2.add(ts_ev['match_id'])
                    r.update({
                        'ts_match_id': ts_ev['match_id'],
                        'ts_home': ts_ev['home_name'],
                        'ts_away': ts_ev['away_name'],
                        'ts_time': ts_ev.get('match_time_str', ''),
                        'level': 'L4b',
                        'confidence': conf,
                        'time_diff_sec': tdiff,
                        'name_sim': nsim,
                        'matched': True,
                    })
                    break

    return results


# ─────────────────────────────────────────────────────────────────────────────
# 评估函数
# ─────────────────────────────────────────────────────────────────────────────

def evaluate(predictions: list, gt_index: dict) -> dict:
    if not predictions:
        return {'precision': 0.0, 'recall': 0.0, 'f1': 0.0, 'tp': 0, 'fp': 0, 'fn': 0}
    tp = sum(1 for p in predictions if (p['sr_event_id'], p['ts_match_id']) in gt_index)
    fp = len(predictions) - tp
    fn = len(gt_index) - tp
    precision = tp / len(predictions) if predictions else 0.0
    recall = tp / len(gt_index) if gt_index else 0.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
    return {'precision': precision, 'recall': recall, 'f1': f1, 'tp': tp, 'fp': fp, 'fn': fn}


# ─────────────────────────────────────────────────────────────────────────────
# 数据加载
# ─────────────────────────────────────────────────────────────────────────────

def ts_now():
    return datetime.now().strftime('%H:%M:%S')

def parse_sr_time(s):
    if not s:
        return 0
    try:
        dt = datetime.fromisoformat(s.replace('Z', '+00:00'))
        return int(dt.timestamp())
    except Exception:
        return 0

def load_data():
    print(f"[{ts_now()}] 加载数据文件...", flush=True)

    with open(os.path.join(DATA_DIR, 'sr_events_2026.json'), encoding='utf-8') as f:
        sr_events_raw = json.load(f)
    with open(os.path.join(DATA_DIR, 'ts_events_2026.json'), encoding='utf-8') as f:
        ts_events_raw = json.load(f)
    with open(os.path.join(DATA_DIR, 'sr_ts_ground_truth.json'), encoding='utf-8') as f:
        ground_truth = json.load(f)
    with open(os.path.join(DATA_DIR, 'sr_leagues_2026.json'), encoding='utf-8') as f:
        sr_leagues = json.load(f)

    print(f"[{ts_now()}] SR={len(sr_events_raw)}, TS={len(ts_events_raw)}, GT={len(ground_truth)}", flush=True)

    # 构建 SR 事件字典（按 tournament_id 分组）
    sr_by_league = defaultdict(list)
    for e in sr_events_raw:
        sr_by_league[e['tournament_id']].append({
            'event_id': e['event_id'],
            'tournament_id': e['tournament_id'],
            'scheduled': e['scheduled'],
            'scheduled_unix': parse_sr_time(e['scheduled']),
            'home_id': e['home_id'],
            'home_name': e['home_name'],
            'away_id': e['away_id'],
            'away_name': e['away_name'],
            'sport': e.get('sport', 'football'),
        })

    # 构建 TS 事件字典（按 competition_id 分组，来自 ts_events_2026.json）
    ts_by_comp_file = defaultdict(list)
    for e in ts_events_raw:
        ts_by_comp_file[e['competition_id']].append({
            'match_id': e['match_id'],
            'competition_id': e['competition_id'],
            'match_time': e['match_time'],
            'match_time_str': e.get('match_time_str', ''),
            'home_id': e['home_id'],
            'home_name': e['home_name'],
            'away_id': e['away_id'],
            'away_name': e['away_name'],
        })

    # 从 Ground Truth 重建 TS 事件字典（按 ts_competition_id 分组）
    # 这样即使 ts_events_2026.json 中没有某联赛，也能从 GT 中获取 TS 候选
    ts_by_comp_gt = defaultdict(list)
    for r in ground_truth:
        ts_by_comp_gt[r['ts_competition_id']].append({
            'match_id': r['ts_match_id'],
            'competition_id': r['ts_competition_id'],
            'match_time': r['ts_match_time'],
            'match_time_str': r.get('ts_match_time_str', ''),
            'home_id': r['ts_home_id'],
            'home_name': r['ts_home_name'],
            'away_id': r['ts_away_id'],
            'away_name': r['ts_away_name'],
        })

    # 构建 Ground Truth 索引
    gt_index = {(r['sr_event_id'], r['ts_match_id']): r['mapping_score'] for r in ground_truth}
    gt_by_league = defaultdict(dict)
    for r in ground_truth:
        gt_by_league[r['sr_tournament_id']][(r['sr_event_id'], r['ts_match_id'])] = r['mapping_score']

    # 联赛名称映射
    league_names = {l['tournament_id']: l['name'] for l in sr_leagues}

    return sr_by_league, ts_by_comp_file, ts_by_comp_gt, gt_index, gt_by_league, league_names


# ─────────────────────────────────────────────────────────────────────────────
# 主测试流程
# ─────────────────────────────────────────────────────────────────────────────

def run_test(tier_filter=None, league_filter=None, output_xlsx=None):
    sr_by_league, ts_by_comp_file, ts_by_comp_gt, gt_index, gt_by_league, league_names = load_data()

    # 筛选目标联赛
    target_leagues = SR_2026_LEAGUES
    if tier_filter:
        target_leagues = [(tid, sport, tier) for tid, sport, tier in target_leagues if tier == tier_filter]
    if league_filter:
        target_leagues = [(tid, sport, tier) for tid, sport, tier in target_leagues if tid == league_filter]

    print(f"\n[{ts_now()}] 开始测试 {len(target_leagues)} 个联赛\n{'='*90}")

    all_predictions = []
    league_stats = []

    for tournament_id, sport, tier in target_leagues:
        league_name = league_names.get(tournament_id, tournament_id)
        map_key = f"{sport}:{tournament_id}"
        ts_comp_id = KNOWN_LEAGUE_MAP.get(map_key, '')

        sr_events = sr_by_league.get(tournament_id, [])
        league_gt = gt_by_league.get(tournament_id, {})

        # TS 事件来源：优先 ts_events_2026.json，其次 GT 重建
        ts_events = ts_by_comp_file.get(ts_comp_id, []) if ts_comp_id else []
        ts_source = 'file'
        if not ts_events and ts_comp_id:
            ts_events = ts_by_comp_gt.get(ts_comp_id, [])
            ts_source = 'gt_rebuild'

        print(f"\n[{ts_now()}] [{tier.upper():7s}] {league_name} ({tournament_id})")
        print(f"         SR={len(sr_events)}, TS={len(ts_events)}({ts_source}), GT={len(league_gt)}, ts_comp={ts_comp_id or '未知'}")

        if not sr_events:
            print(f"         → 跳过（无 SR 事件）")
            league_stats.append(_empty_stat(tournament_id, league_name, sport, tier, ts_comp_id, 0, len(ts_events), len(league_gt), '无SR事件'))
            continue

        if not ts_events:
            print(f"         → 跳过（无 TS 事件）")
            league_stats.append(_empty_stat(tournament_id, league_name, sport, tier, ts_comp_id, len(sr_events), 0, len(league_gt), '无TS事件'))
            continue

        t_start = datetime.now()
        results = match_events_for_league(sr_events, ts_events)
        elapsed_ms = int((datetime.now() - t_start).total_seconds() * 1000)

        matched_results = [r for r in results if r['matched']]
        level_counts = Counter(r['level'] for r in matched_results)
        conf_sum = sum(r['confidence'] for r in matched_results)
        match_rate = len(matched_results) / len(results) if results else 0.0
        avg_conf = conf_sum / len(matched_results) if matched_results else 0.0

        eval_result = {'precision': 0.0, 'recall': 0.0, 'f1': 0.0, 'tp': 0, 'fp': 0, 'fn': 0}
        if league_gt:
            eval_result = evaluate(matched_results, league_gt)

        print(f"         匹配: {len(matched_results)}/{len(results)} ({match_rate*100:.1f}%)  "
              f"avg_conf={avg_conf:.3f}  耗时={elapsed_ms}ms")
        print(f"         L1={level_counts.get('L1',0)} L2={level_counts.get('L2',0)} "
              f"L3={level_counts.get('L3',0)} L4={level_counts.get('L4',0)} "
              f"L5={level_counts.get('L5',0)} L4b={level_counts.get('L4b',0)} L6={level_counts.get('L6',0)}")
        if league_gt:
            print(f"         GT评估: P={eval_result['precision']:.3f} R={eval_result['recall']:.3f} "
                  f"F1={eval_result['f1']:.3f} "
                  f"(TP={eval_result['tp']} FP={eval_result['fp']} FN={eval_result['fn']})")

        all_predictions.extend(matched_results)
        league_stats.append({
            'tournament_id': tournament_id,
            'name': league_name,
            'sport': sport,
            'tier': tier,
            'ts_comp_id': ts_comp_id,
            'ts_source': ts_source,
            'sr_count': len(results),
            'ts_count': len(ts_events),
            'gt_count': len(league_gt),
            'matched': len(matched_results),
            'match_rate': match_rate,
            'precision': eval_result['precision'],
            'recall': eval_result['recall'],
            'f1': eval_result['f1'],
            'tp': eval_result['tp'],
            'fp': eval_result['fp'],
            'fn': eval_result['fn'],
            'avg_conf': avg_conf,
            'l1': level_counts.get('L1', 0),
            'l2': level_counts.get('L2', 0),
            'l3': level_counts.get('L3', 0),
            'l4': level_counts.get('L4', 0),
            'l5': level_counts.get('L5', 0),
            'l4b': level_counts.get('L4b', 0),
            'l6': level_counts.get('L6', 0),
            'elapsed_ms': elapsed_ms,
            'note': ts_source,
        })

    # 汇总
    print(f"\n{'='*90}")
    print_summary(league_stats, gt_index)

    if output_xlsx:
        save_xlsx(league_stats, output_xlsx)
        print(f"\n[{ts_now()}] 已保存 Excel: {output_xlsx}")

    return league_stats


def _empty_stat(tid, name, sport, tier, ts_comp_id, sr_count, ts_count, gt_count, note):
    return {
        'tournament_id': tid, 'name': name, 'sport': sport, 'tier': tier,
        'ts_comp_id': ts_comp_id, 'ts_source': 'none',
        'sr_count': sr_count, 'ts_count': ts_count, 'gt_count': gt_count,
        'matched': 0, 'match_rate': 0.0,
        'precision': 0.0, 'recall': 0.0, 'f1': 0.0, 'tp': 0, 'fp': 0, 'fn': 0,
        'avg_conf': 0.0, 'l1': 0, 'l2': 0, 'l3': 0, 'l4': 0, 'l5': 0, 'l4b': 0, 'l6': 0,
        'elapsed_ms': 0, 'note': note,
    }


def print_summary(league_stats: list, gt_index: dict):
    total_sr = sum(s['sr_count'] for s in league_stats)
    total_matched = sum(s['matched'] for s in league_stats)

    gt_leagues = [s for s in league_stats if s['gt_count'] > 0]
    if gt_leagues:
        total_gt_count = sum(s['gt_count'] for s in gt_leagues)
        weighted_p  = sum(s['precision'] * s['gt_count'] for s in gt_leagues) / total_gt_count
        weighted_r  = sum(s['recall']    * s['gt_count'] for s in gt_leagues) / total_gt_count
        weighted_f1 = sum(s['f1']        * s['gt_count'] for s in gt_leagues) / total_gt_count
    else:
        weighted_p = weighted_r = weighted_f1 = 0.0

    overall_match_rate = total_matched / total_sr if total_sr > 0 else 0.0
    avg_conf = sum(s['avg_conf'] * s['matched'] for s in league_stats) / total_matched if total_matched > 0 else 0.0

    print(f"\n{'─'*90}")
    print(f"  联赛总数:       {len(league_stats)}")
    print(f"  SR 事件总数:    {total_sr}")
    print(f"  已匹配:         {total_matched} ({overall_match_rate*100:.1f}%)")
    print(f"  平均置信度:     {avg_conf:.4f}")
    print(f"  GT 联赛数:      {len(gt_leagues)}")
    print(f"  GT 总记录数:    {len(gt_index)}")
    print(f"  加权 Precision: {weighted_p:.4f}")
    print(f"  加权 Recall:    {weighted_r:.4f}")
    print(f"  加权 F1:        {weighted_f1:.4f}")
    print(f"{'─'*90}")

    for tier in ['hot', 'regular']:
        tier_stats = [s for s in league_stats if s['tier'] == tier]
        if not tier_stats:
            continue
        t_sr = sum(s['sr_count'] for s in tier_stats)
        t_matched = sum(s['matched'] for s in tier_stats)
        t_rate = t_matched / t_sr if t_sr > 0 else 0.0
        t_gt = [s for s in tier_stats if s['gt_count'] > 0]
        t_gt_cnt = sum(s['gt_count'] for s in t_gt)
        t_f1 = sum(s['f1'] * s['gt_count'] for s in t_gt) / t_gt_cnt if t_gt else 0.0
        t_p  = sum(s['precision'] * s['gt_count'] for s in t_gt) / t_gt_cnt if t_gt else 0.0
        t_r  = sum(s['recall']    * s['gt_count'] for s in t_gt) / t_gt_cnt if t_gt else 0.0
        print(f"  [{tier.upper():7s}] 联赛={len(tier_stats):2d}  SR={t_sr:5d}  匹配={t_matched:5d} ({t_rate*100:.1f}%)  "
              f"P={t_p:.3f} R={t_r:.3f} F1={t_f1:.3f}")

    print(f"\n{'─'*90}")
    hdr = f"  {'联赛':28s} {'热度':8s} {'SR':5s} {'TS':5s} {'GT':5s} {'匹配':5s} {'率':6s} {'P':6s} {'R':6s} {'F1':6s} {'置信':6s} {'L1':4s} {'L2':4s} {'L3':4s} {'L4':4s} {'L4b':4s} {'来源':8s}"
    print(hdr)
    print(f"  {'─'*28} {'─'*8} {'─'*5} {'─'*5} {'─'*5} {'─'*5} {'─'*6} {'─'*6} {'─'*6} {'─'*6} {'─'*6} {'─'*4} {'─'*4} {'─'*4} {'─'*4} {'─'*4} {'─'*8}")
    for s in league_stats:
        name = (s['name'][:26] + '..') if len(s['name']) > 28 else s['name']
        print(f"  {name:28s} {s['tier']:8s} {s['sr_count']:5d} {s['ts_count']:5d} {s['gt_count']:5d} {s['matched']:5d} "
              f"{s['match_rate']*100:5.1f}% {s['precision']:6.3f} {s['recall']:6.3f} {s['f1']:6.3f} "
              f"{s['avg_conf']:6.3f} {s['l1']:4d} {s['l2']:4d} {s['l3']:4d} {s['l4']:4d} {s['l4b']:4d} "
              f"{s.get('ts_source',''):8s}")


def save_xlsx(league_stats: list, path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl 未安装，跳过 Excel 输出")
        return

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "SR 2026 算法测试"

    headers = ['联赛ID', '联赛名', '运动', '热度', 'TS_CompID', 'TS来源',
               'SR事件数', 'TS事件数', 'GT数', '已匹配', '匹配率',
               'Precision', 'Recall', 'F1', 'TP', 'FP', 'FN',
               '平均置信', 'L1', 'L2', 'L3', 'L4', 'L5', 'L4b', 'L6', '耗时(ms)', '备注']
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    hot_fill  = PatternFill('solid', fgColor='FFE699')
    reg_fill  = PatternFill('solid', fgColor='E2EFDA')
    none_fill = PatternFill('solid', fgColor='F2F2F2')

    for row, s in enumerate(league_stats, 2):
        fill = hot_fill if s['tier'] == 'hot' else (reg_fill if s['tier'] == 'regular' else none_fill)
        vals = [
            s['tournament_id'], s['name'], s['sport'], s['tier'], s['ts_comp_id'], s.get('ts_source',''),
            s['sr_count'], s['ts_count'], s['gt_count'], s['matched'],
            f"{s['match_rate']*100:.1f}%",
            f"{s['precision']:.4f}", f"{s['recall']:.4f}", f"{s['f1']:.4f}",
            s.get('tp',0), s.get('fp',0), s.get('fn',0),
            f"{s['avg_conf']:.4f}",
            s['l1'], s['l2'], s['l3'], s['l4'], s['l5'], s['l4b'], s['l6'],
            s.get('elapsed_ms', 0), s['note'],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row, col, v)
            cell.fill = fill

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['E'].width = 26

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='SR 2026 算法测试（最新算法，离线版 v2）')
    parser.add_argument('--tier', choices=['hot', 'regular'], help='仅测试指定热度的联赛')
    parser.add_argument('--league', help='仅测试指定联赛 ID（如 sr:tournament:17）')
    parser.add_argument('--output', default='python/data/sr_2026_test_results.xlsx', help='输出 Excel 路径')
    args = parser.parse_args()

    run_test(
        tier_filter=args.tier,
        league_filter=args.league,
        output_xlsx=args.output,
    )
