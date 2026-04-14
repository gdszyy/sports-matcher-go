#!/usr/bin/env python3
"""
LSports ↔ TheSports 2026年比赛匹配脚本 v3（最终优化版）
优化策略：
- 一次性预加载所有2026年TS比赛（避免76次单独查询）
- 预加载球队名称字典（避免JOIN）
- 在内存中按competition_id分组，O(1)查找
- 时间窗口索引加速比赛匹配
"""

import pymysql
import difflib
import unicodedata
import re
import bisect
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
import calendar

# ─────────────────────────────────────────────────────────────────────────────
# 配置
# ─────────────────────────────────────────────────────────────────────────────
DB_HOST = '127.0.0.1'
DB_PORT = 3308
DB_PASSWORD = 'r74pqyYtgdjlYB41jmWA'
LS_DB = 'test-xp-lsports'
TS_DB = 'test-thesports-db'

YEAR_2026_START = '2026-01-01'
YEAR_2026_END = '2027-01-01'
TS_2026_START = 1767225600  # 2026-01-01 00:00:00 UTC
TS_2026_END = 1798761600    # 2027-01-01 00:00:00 UTC

def ts(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def get_conn(db):
    return pymysql.connect(
        host=DB_HOST, port=DB_PORT,
        user='root', password=DB_PASSWORD,
        database=db, charset='utf8mb4', connect_timeout=20
    )

# ─────────────────────────────────────────────────────────────────────────────
# 已知联赛映射
# ─────────────────────────────────────────────────────────────────────────────
KNOWN_LS_TS_MAP = {
    "football:67":    "jednm9whz0ryox8",   # Premier League (England)
    "football:8363":  "vl7oqdehlyr510j",   # LaLiga (Spain)
    "football:61":    "yl5ergphnzr8k0o",   # Ligue 1 (France)
    "football:65":    "gy0or5jhg6qwzv3",   # Bundesliga (Germany)
    "football:66":    "kn54qllhjzqvy9d",   # 2.Bundesliga (Germany)
    "football:32644": "z8yomo4h7wq0j6l",   # UEFA Champions League
    "football:30444": "56ypq3nh0xmd7oj",   # UEFA Europa League
    "basketball:64":  "49vjxm8xt4q6odg",   # NBA
}

COUNTRY_NAME_MAP = {
    "chinese": "China", "china": "China",
    "japanese": "Japan", "japan": "Japan",
    "korean": "Korea", "korea": "Korea",
    "south korean": "South Korea", "south korea": "South Korea",
    "indian": "India", "india": "India",
    "thai": "Thailand", "thailand": "Thailand",
    "vietnamese": "Vietnam", "vietnam": "Vietnam",
    "indonesian": "Indonesia", "indonesia": "Indonesia",
    "malaysian": "Malaysia", "malaysia": "Malaysia",
    "singaporean": "Singapore", "singapore": "Singapore",
    "philippine": "Philippines", "philippines": "Philippines",
    "iranian": "Iran", "iran": "Iran",
    "saudi": "Saudi Arabia", "saudi arabia": "Saudi Arabia",
    "qatari": "Qatar", "qatar": "Qatar",
    "emirati": "UAE", "uae": "UAE", "united arab emirates": "UAE",
    "english": "England", "england": "England",
    "scottish": "Scotland", "scotland": "Scotland",
    "welsh": "Wales", "wales": "Wales",
    "german": "Germany", "germany": "Germany",
    "french": "France", "france": "France",
    "spanish": "Spain", "spain": "Spain",
    "italian": "Italy", "italy": "Italy",
    "portuguese": "Portugal", "portugal": "Portugal",
    "dutch": "Netherlands", "netherlands": "Netherlands", "holland": "Netherlands",
    "belgian": "Belgium", "belgium": "Belgium",
    "swiss": "Switzerland", "switzerland": "Switzerland",
    "austrian": "Austria", "austria": "Austria",
    "swedish": "Sweden", "sweden": "Sweden",
    "norwegian": "Norway", "norway": "Norway",
    "danish": "Denmark", "denmark": "Denmark",
    "finnish": "Finland", "finland": "Finland",
    "russian": "Russia", "russia": "Russia",
    "ukrainian": "Ukraine", "ukraine": "Ukraine",
    "polish": "Poland", "poland": "Poland",
    "czech": "Czech Republic", "czechia": "Czech Republic",
    "hungarian": "Hungary", "hungary": "Hungary",
    "romanian": "Romania", "romania": "Romania",
    "greek": "Greece", "greece": "Greece",
    "turkish": "Turkey", "turkey": "Turkey",
    "croatian": "Croatia", "croatia": "Croatia",
    "serbian": "Serbia", "serbia": "Serbia",
    "american": "USA", "usa": "USA", "united states": "USA",
    "canadian": "Canada", "canada": "Canada",
    "mexican": "Mexico", "mexico": "Mexico",
    "brazilian": "Brazil", "brazil": "Brazil",
    "argentine": "Argentina", "argentina": "Argentina",
    "colombian": "Colombia", "colombia": "Colombia",
    "chilean": "Chile", "chile": "Chile",
    "australian": "Australia", "australia": "Australia",
    "south african": "South Africa", "south africa": "South Africa",
    "nigerian": "Nigeria", "nigeria": "Nigeria",
    "moroccan": "Morocco", "morocco": "Morocco",
    "egyptian": "Egypt", "egypt": "Egypt",
    "hong kong": "Hong Kong",
    "taiwan": "Taiwan",
}

INTERNATIONAL_KEYWORDS = {
    'world', 'international', 'europe', 'europa', 'asia', 'africa',
    'america', 'oceania', 'concacaf', 'conmebol', 'afc', 'caf',
    'uefa', 'fifa', 'ofc', 'waff', 'saff', 'eaff',
}

# ─────────────────────────────────────────────────────────────────────────────
# 名称归一化与相似度
# ─────────────────────────────────────────────────────────────────────────────

_norm_cache = {}

def normalize_name(s: str) -> str:
    if s in _norm_cache:
        return _norm_cache[s]
    if not s:
        _norm_cache[s] = ''
        return ''
    r = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    r = r.lower()
    r = re.sub(r'[.\-_,\'\"·|]', ' ', r)
    r = ' '.join(r.split())
    _norm_cache[s] = r
    return r

def name_similarity(a: str, b: str) -> float:
    na, nb = normalize_name(a), normalize_name(b)
    if not na and not nb:
        return 1.0
    if not na or not nb:
        return 0.0
    # Jaccard
    sa, sb = set(na.split()), set(nb.split())
    j = len(sa & sb) / len(sa | sb) if (sa | sb) else 0.0
    # Sequence
    s = difflib.SequenceMatcher(None, na, nb).ratio()
    return max(j, s)

def is_international(name: str) -> bool:
    if not name:
        return False
    tokens = set(normalize_name(name).split())
    return bool(tokens & INTERNATIONAL_KEYWORDS)

def extract_country(name: str) -> str:
    if not name:
        return ''
    tokens = normalize_name(name).split()
    for length in (3, 2, 1):
        if len(tokens) >= length:
            phrase = ' '.join(tokens[:length])
            if phrase in COUNTRY_NAME_MAP:
                return COUNTRY_NAME_MAP[phrase]
    return ''

_eff_country_cache = {}

def get_eff_country(comp: dict) -> str:
    cid = comp.get('competition_id', '')
    if cid in _eff_country_cache:
        return _eff_country_cache[cid]
    c = comp.get('host_country', '') or ''
    if not c:
        c = extract_country(comp.get('name', ''))
    _eff_country_cache[cid] = c
    return c

# ─────────────────────────────────────────────────────────────────────────────
# 联赛匹配
# ─────────────────────────────────────────────────────────────────────────────

def match_league(ls_name, ls_category, ts_comps, sport, ls_id) -> dict:
    empty = {'ts_id': '', 'ts_name': '', 'ts_country': '', 'ts_eff_country': '',
             'score': 0.0, 'rule': 'NO_MATCH', 'matched': False}

    # 已知映射
    key = f"{sport}:{ls_id}"
    if key in KNOWN_LS_TS_MAP:
        ts_id = KNOWN_LS_TS_MAP[key]
        for c in ts_comps:
            if c['competition_id'] == ts_id:
                return {'ts_id': ts_id, 'ts_name': c['name'],
                        'ts_country': c.get('host_country', ''),
                        'ts_eff_country': get_eff_country(c),
                        'score': 1.0, 'rule': 'KNOWN', 'matched': True}
        return {'ts_id': ts_id, 'ts_name': '', 'ts_country': '', 'ts_eff_country': '',
                'score': 1.0, 'rule': 'KNOWN', 'matched': True}

    best_score = 0.0
    best_comp = None
    best_eff = ''

    for c in ts_comps:
        eff = get_eff_country(c)
        # 地理否决
        if ls_category and eff and not is_international(ls_category) and not is_international(eff):
            if name_similarity(ls_category, eff) < 0.4:
                continue
        base = name_similarity(ls_name, c['name'])
        if ls_category and eff:
            loc = name_similarity(ls_category, eff)
            if loc >= 0.6:
                base = base * 0.75 + 0.25 * loc
            elif loc >= 0.4:
                base = base * (0.70 + 0.30 * (loc - 0.4) / 0.2)
        if base > best_score:
            best_score = base
            best_comp = c
            best_eff = eff

    if best_comp is None or best_score < 0.55:
        return empty

    rule = 'NAME_HI' if best_score >= 0.85 else ('NAME_MED' if best_score >= 0.70 else 'NAME_LOW')
    return {'ts_id': best_comp['competition_id'], 'ts_name': best_comp['name'],
            'ts_country': best_comp.get('host_country', ''), 'ts_eff_country': best_eff,
            'score': round(best_score, 4), 'rule': rule, 'matched': True}

# ─────────────────────────────────────────────────────────────────────────────
# 比赛匹配（时间窗口索引）
# ─────────────────────────────────────────────────────────────────────────────

def parse_ls_time(s: str) -> int:
    if not s:
        return 0
    s = s.strip().rstrip('Z')
    if s.endswith('+00:00'):
        s = s[:-6]
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return int(calendar.timegm(datetime.strptime(s, fmt).timetuple()))
        except:
            pass
    return 0

def team_sim(lh, la, th, ta):
    fwd = (name_similarity(lh, th) + name_similarity(la, ta)) / 2
    rev = (name_similarity(lh, ta) + name_similarity(la, th)) / 2
    return max(fwd, rev)

def same_date(t1, t2):
    if not t1 or not t2:
        return False
    return (datetime.fromtimestamp(t1, tz=timezone.utc).date() ==
            datetime.fromtimestamp(t2, tz=timezone.utc).date())

def match_events(ls_events, ts_events):
    if not ts_events:
        return [_no_match(e) for e in ls_events]

    ts_sorted = sorted(ts_events, key=lambda x: x['match_time'] or 0)
    ts_times = [e['match_time'] or 0 for e in ts_sorted]
    ts_used = set()

    LEVELS = [
        ('L1', 300, 0.40),
        ('L2', 21600, 0.65),
        ('L3', 86400, 0.75),
        ('L4', 259200, 0.85),
    ]

    assignments = {}

    for level, window, min_score in LEVELS:
        for ev in ls_events:
            if ev['event_id'] in assignments:
                continue
            lt = ev['start_unix']
            lo = bisect.bisect_left(ts_times, lt - window)
            hi = bisect.bisect_right(ts_times, lt + window)
            candidates = ts_sorted[lo:hi]

            best_score = 0.0
            best_ts = None

            for te in candidates:
                if te['match_id'] in ts_used:
                    continue
                tt = te['match_time'] or 0
                diff = abs(lt - tt) if lt and tt else 999999

                if level == 'L1' and diff > 300:
                    continue
                elif level == 'L2' and diff > 21600:
                    continue
                elif level == 'L3' and not same_date(lt, tt):
                    continue
                elif level == 'L4' and diff > 259200:
                    continue

                sc = team_sim(ev['home_name'], ev['away_name'],
                              te['home_name'], te['away_name'])
                if sc >= min_score and sc > best_score:
                    best_score = sc
                    best_ts = te

            if best_ts:
                assignments[ev['event_id']] = (best_ts, best_score, level)
                ts_used.add(best_ts['match_id'])

    results = []
    for ev in ls_events:
        if ev['event_id'] in assignments:
            te, score, level = assignments[ev['event_id']]
            ts_time_str = (datetime.utcfromtimestamp(te['match_time']).strftime('%Y-%m-%dT%H:%M:%S')
                           if te['match_time'] else '')
            results.append({
                'ls_event_id': ev['event_id'],
                'ls_scheduled': ev['scheduled'],
                'ls_home_id': ev['home_id'],
                'ls_home_name': ev['home_name'],
                'ls_away_id': ev['away_id'],
                'ls_away_name': ev['away_name'],
                'ts_match_id': te['match_id'],
                'ts_match_time': te['match_time'],
                'ts_match_time_str': ts_time_str,
                'ts_home_id': te['home_team_id'],
                'ts_home_name': te['home_name'],
                'ts_away_id': te['away_team_id'],
                'ts_away_name': te['away_name'],
                'team_score': round(score, 4),
                'match_level': level,
                'matched': True,
            })
        else:
            results.append(_no_match(ev))
    return results

def _no_match(ev):
    return {
        'ls_event_id': ev['event_id'], 'ls_scheduled': ev['scheduled'],
        'ls_home_id': ev['home_id'], 'ls_home_name': ev['home_name'],
        'ls_away_id': ev['away_id'], 'ls_away_name': ev['away_name'],
        'ts_match_id': '', 'ts_match_time': 0, 'ts_match_time_str': '',
        'ts_home_id': '', 'ts_home_name': '', 'ts_away_id': '', 'ts_away_name': '',
        'team_score': 0.0, 'match_level': 'NO_MATCH', 'matched': False,
    }

# ─────────────────────────────────────────────────────────────────────────────
# 数据加载（批量优化）
# ─────────────────────────────────────────────────────────────────────────────

def load_all_ts_data(sport):
    """一次性加载所有2026年TS比赛，按competition_id分组"""
    conn = get_conn(TS_DB)
    with conn.cursor() as cur:
        # 球队名称字典
        if sport == 'football':
            cur.execute("SELECT team_id, name FROM ts_fb_team")
        else:
            cur.execute("SELECT team_id, name FROM ts_bb_team")
        teams = {str(r[0]): r[1] for r in cur.fetchall()}  # key统一为str
        ts(f"  TS {sport} teams loaded: {len(teams)}")

        # 所有2026年比赛
        if sport == 'football':
            cur.execute("""
                SELECT match_id, competition_id, home_team_id, away_team_id, match_time
                FROM ts_fb_match
                WHERE match_time >= %s AND match_time < %s
                ORDER BY match_time
            """, (TS_2026_START, TS_2026_END))
        else:
            cur.execute("""
                SELECT match_id, competition_id, home_team_id, away_team_id, match_time
                FROM ts_bb_match
                WHERE match_time >= %s AND match_time < %s
                ORDER BY match_time
            """, (TS_2026_START, TS_2026_END))
        rows = cur.fetchall()
        ts(f"  TS {sport} 2026 matches loaded: {len(rows)}")
    conn.close()

    by_comp = defaultdict(list)
    for row in rows:
        by_comp[row[1]].append({
            'match_id': row[0],
            'competition_id': row[1],
            'home_team_id': row[2],
            'home_name': teams.get(str(row[2]), ''),  # str key
            'away_team_id': row[3],
            'away_name': teams.get(str(row[3]), ''),  # str key
            'match_time': row[4],
        })
    return by_comp


def load_ls_tournaments_2026(sport):
    sport_id = '6046' if sport == 'football' else '48242'
    conn = get_conn(LS_DB)
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        # 使用COUNT(DISTINCT)避免JOIN放大
        cur.execute("""
            SELECT e.tournament_id,
                   COALESCE(t.name, '') as tournament_name,
                   COALESCE(cat.name, '') as location,
                   COUNT(DISTINCT e.event_id) as event_count
            FROM ls_sport_event e
            LEFT JOIN ls_tournament_en t ON e.tournament_id = t.tournament_id
            LEFT JOIN ls_category_en cat ON t.category_id = cat.category_id
                      AND cat.sport_id = e.sport_id
            WHERE e.sport_id = %s
              AND e.scheduled LIKE '2026%%'
            GROUP BY e.tournament_id, t.name, cat.name
            ORDER BY event_count DESC
        """, (sport_id,))
        rows = cur.fetchall()
    conn.close()
    return rows


def load_ls_events_2026_bulk(sport):
    """一次性加载所有2026年LS比赛"""
    sport_id = '6046' if sport == 'football' else '48242'
    conn = get_conn(LS_DB)
    with conn.cursor() as cur:
        # 球队名称（competitor_id是bigint，但event表中是varchar，统一转str）
        cur.execute("SELECT competitor_id, name FROM ls_competitor_en")
        competitors = {str(r[0]): r[1] for r in cur.fetchall()}
        ts(f"  LS competitors loaded: {len(competitors)}")

        # 所有2026年比赛
        cur.execute("""
            SELECT event_id, tournament_id, scheduled,
                   home_competitor_id, away_competitor_id, status
            FROM ls_sport_event
            WHERE sport_id = %s
              AND scheduled LIKE '2026%%'
            ORDER BY scheduled
        """, (sport_id,))
        rows = cur.fetchall()
        ts(f"  LS {sport} 2026 events loaded: {len(rows)}")
    conn.close()

    by_tournament = defaultdict(list)
    seen = set()
    for row in rows:
        eid = row[0]
        if eid in seen:
            continue
        seen.add(eid)
        ev = {
            'event_id': eid,
            'tournament_id': str(row[1]),
            'scheduled': row[2] or '',
            'home_id': row[3] or '',
            'home_name': competitors.get(str(row[3]), '') if row[3] else '',  # str key
            'away_id': row[4] or '',
            'away_name': competitors.get(str(row[4]), '') if row[4] else '',  # str key
            'status_id': row[5] or 0,
        }
        ev['start_unix'] = parse_ls_time(ev['scheduled'])
        by_tournament[str(row[1])].append(ev)
    return by_tournament


def load_ts_competitions(sport):
    conn = get_conn(TS_DB)
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        if sport == 'football':
            cur.execute("SELECT competition_id, name, COALESCE(host_country,'') as host_country FROM ts_fb_competition")
        else:
            cur.execute("SELECT competition_id, name, '' as host_country FROM ts_bb_competition")
        rows = cur.fetchall()
    conn.close()
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────────────────────────

def run_sport_match(sport):
    _eff_country_cache.clear()
    _norm_cache.clear()

    ts(f"\n{'='*60}")
    ts(f"开始 {sport} 匹配...")
    ts(f"{'='*60}")

    # 加载数据
    ts("加载 LSports 联赛...")
    ls_tours = load_ls_tournaments_2026(sport)
    ts(f"  LSports 联赛: {len(ls_tours)}")

    ts("加载 LSports 比赛（批量）...")
    ls_events_by_tour = load_ls_events_2026_bulk(sport)

    ts("加载 TheSports 联赛...")
    ts_comps = load_ts_competitions(sport)
    ts(f"  TheSports 联赛: {len(ts_comps)}")

    ts("加载 TheSports 比赛（批量）...")
    ts_events_by_comp = load_all_ts_data(sport)

    # 预计算TS联赛国家
    for c in ts_comps:
        get_eff_country(c)

    # 联赛匹配
    ts("执行联赛匹配...")
    league_results = []
    matched_leagues = 0

    for tour in ls_tours:
        ls_id = str(tour['tournament_id'])
        ls_name = tour['tournament_name'] or ''
        ls_category = tour['location'] or ''
        ls_event_count = tour['event_count']

        m = match_league(ls_name, ls_category, ts_comps, sport, ls_id)
        if m['matched']:
            matched_leagues += 1

        league_results.append({
            'ls_tournament_id': ls_id,
            'ls_name': ls_name,
            'ls_category': ls_category,
            'ls_event_count_2026': ls_event_count,
            'ts_competition_id': m['ts_id'],
            'ts_name': m['ts_name'],
            'ts_country': m['ts_country'],
            'ts_eff_country': m['ts_eff_country'],
            'league_score': m['score'],
            'league_rule': m['rule'],
            'league_matched': m['matched'],
        })

    ts(f"  联赛匹配: {matched_leagues}/{len(ls_tours)}")

    # 比赛匹配
    ts("执行比赛匹配...")
    league_event_results = {}
    total_ls = 0
    total_matched = 0

    for lr in league_results:
        ls_id = lr['ls_tournament_id']
        ls_events = ls_events_by_tour.get(ls_id, [])

        if not lr['league_matched']:
            lr['ts_event_count_2026'] = 0
            lr['event_matched_count'] = 0
            lr['event_match_rate'] = 'N/A'
            continue

        ts_comp_id = lr['ts_competition_id']
        ts_events = ts_events_by_comp.get(ts_comp_id, [])

        if not ls_events:
            lr['ts_event_count_2026'] = len(ts_events)
            lr['event_matched_count'] = 0
            lr['event_match_rate'] = 'N/A'
            continue

        ts(f"  {lr['ls_name']} ({ls_id}): LS={len(ls_events)}, TS={len(ts_events)}")

        event_matches = match_events(ls_events, ts_events)
        matched_count = sum(1 for e in event_matches if e['matched'])

        total_ls += len(ls_events)
        total_matched += matched_count

        lr['ts_event_count_2026'] = len(ts_events)
        lr['event_matched_count'] = matched_count
        lr['event_match_rate'] = f"{matched_count/len(ls_events)*100:.1f}%"

        league_event_results[ls_id] = {
            'league_info': lr,
            'event_matches': event_matches,
            'ls_event_count': len(ls_events),
            'ts_event_count': len(ts_events),
            'matched_count': matched_count,
        }

    ts(f"\n{sport} 完成: 联赛 {matched_leagues}/{len(ls_tours)}, 比赛 {total_matched}/{total_ls}")

    return {
        'sport': sport,
        'league_results': league_results,
        'league_event_results': league_event_results,
        'stats': {
            'total_leagues': len(ls_tours),
            'matched_leagues': matched_leagues,
            'total_ls_events': total_ls,
            'total_matched_events': total_matched,
        }
    }

# ─────────────────────────────────────────────────────────────────────────────
# Excel 导出
# ─────────────────────────────────────────────────────────────────────────────

C_BLUE = '1F4E79'
C_LIGHT = '2E75B6'
C_KNOWN = 'C6EFCE'
C_HI = 'DDEEFF'
C_MED = 'FFEB9C'
C_LOW = 'FFD7B5'
C_NO = 'FFC7CE'
C_L1 = 'C6EFCE'
C_L2 = 'DDEEFF'
C_L3 = 'FFEB9C'
C_L4 = 'FFD7B5'

def fill(c):
    return PatternFill(start_color=c, end_color=c, fill_type='solid')

def hdr_font(bold=True):
    return Font(color='FFFFFF', bold=bold)

def auto_width(ws, mn=8, mx=50):
    for col in ws.columns:
        ml = 0
        cl = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    ml = max(ml, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[cl].width = min(max(ml + 2, mn), mx)


def export_excel(fb_data, bb_data, path):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = '联赛匹配统计'

    # 大标题
    ws0.merge_cells('A1:O1')
    ws0['A1'].value = f'LSports → TheSports 联赛匹配统计（2026年）  生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws0['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws0['A1'].fill = fill(C_BLUE)
    ws0['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws0.row_dimensions[1].height = 28

    # 汇总统计
    ws0.append([])
    ws0.append(['运动类型', '总联赛数', '匹配联赛数', '联赛匹配率',
                'LS总比赛数(2026)', '已匹配比赛数', '比赛匹配率'])
    for cell in ws0[3]:
        cell.fill = fill(C_LIGHT)
        cell.font = hdr_font()
        cell.alignment = Alignment(horizontal='center')

    for data in [fb_data, bb_data]:
        s = data['stats']
        sn = '足球 (Football)' if data['sport'] == 'football' else '篮球 (Basketball)'
        lr = f"{s['matched_leagues']/s['total_leagues']*100:.1f}%" if s['total_leagues'] else '0%'
        er = f"{s['total_matched_events']/s['total_ls_events']*100:.1f}%" if s['total_ls_events'] else 'N/A'
        ws0.append([sn, s['total_leagues'], s['matched_leagues'], lr,
                    s['total_ls_events'], s['total_matched_events'], er])
        for cell in ws0[ws0.max_row]:
            cell.alignment = Alignment(horizontal='center')

    ws0.append([])

    # 颜色图例
    ws0.append(['颜色图例：'])
    ws0['A' + str(ws0.max_row)].font = Font(bold=True)
    for color, desc in [
        (C_KNOWN, 'KNOWN - 已知映射'),
        (C_HI, 'NAME_HI - 高置信度名称匹配（≥0.85）'),
        (C_MED, 'NAME_MED - 中置信度名称匹配（≥0.70）'),
        (C_LOW, 'NAME_LOW - 低置信度名称匹配（≥0.55）'),
        (C_NO, 'NO_MATCH - 未匹配'),
    ]:
        ws0.append(['', desc])
        ws0.cell(ws0.max_row, 2).fill = fill(color)

    ws0.append([])

    # 联赛详细列表
    dhr = ws0.max_row + 1
    ws0.append([
        '运动类型',
        'LS联赛ID', 'LS联赛名称', 'LS地区/国家', 'LS 2026比赛数',
        'TS联赛ID', 'TS联赛名称', 'TS国家(DB)', 'TS国家(提取)', 'TS 2026比赛数',
        '联赛相似分', '联赛匹配规则', '联赛已匹配',
        '比赛已匹配数', '比赛匹配率',
    ])
    for cell in ws0[dhr]:
        cell.fill = fill(C_LIGHT)
        cell.font = hdr_font()
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws0.row_dimensions[dhr].height = 30

    rule_fill = {
        'KNOWN': fill(C_KNOWN), 'NAME_HI': fill(C_HI),
        'NAME_MED': fill(C_MED), 'NAME_LOW': fill(C_LOW), 'NO_MATCH': fill(C_NO),
    }

    for data in [fb_data, bb_data]:
        sn = '足球' if data['sport'] == 'football' else '篮球'
        for lr in data['league_results']:
            ws0.append([
                sn,
                lr['ls_tournament_id'], lr['ls_name'], lr['ls_category'], lr['ls_event_count_2026'],
                lr['ts_competition_id'], lr['ts_name'], lr['ts_country'], lr['ts_eff_country'],
                lr.get('ts_event_count_2026', 0),
                lr['league_score'], lr['league_rule'], 'YES' if lr['league_matched'] else 'NO',
                lr.get('event_matched_count', 0), lr.get('event_match_rate', 'N/A'),
            ])
            f = rule_fill.get(lr['league_rule'], fill(C_NO))
            for cell in ws0[ws0.max_row]:
                cell.fill = f
                cell.alignment = Alignment(horizontal='center')
            ws0.cell(ws0.max_row, 3).alignment = Alignment(horizontal='left')
            ws0.cell(ws0.max_row, 7).alignment = Alignment(horizontal='left')

    auto_width(ws0)

    # 每联赛一个Sheet
    ev_fill = {'L1': fill(C_L1), 'L2': fill(C_L2), 'L3': fill(C_L3),
               'L4': fill(C_L4), 'NO_MATCH': fill(C_NO)}

    for data in [fb_data, bb_data]:
        sport = data['sport']
        pfx = 'FB' if sport == 'football' else 'BB'
        sport_cn = '足球' if sport == 'football' else '篮球'

        for ls_id, ld in data['league_event_results'].items():
            lr = ld['league_info']
            ems = ld['event_matches']

            # Sheet名（去除非法字符，限31字符）
            sname = re.sub(r'[\\/*?:\[\]|]', '_', lr['ls_name'])[:18] if lr['ls_name'] else ls_id
            sheet_name = f"{pfx}_{ls_id}_{sname}"[:31]

            ws = wb.create_sheet(title=sheet_name)

            # 标题
            ws.merge_cells('A1:O1')
            ws['A1'].value = (f"【{sport_cn}】{lr['ls_name']}  →  {lr['ts_name']}  "
                              f"| 规则:{lr['league_rule']}  "
                              f"| 比赛匹配:{lr.get('event_matched_count',0)}/{lr['ls_event_count_2026']}  "
                              f"| 匹配率:{lr.get('event_match_rate','N/A')}")
            ws['A1'].font = Font(bold=True, size=11, color='FFFFFF')
            ws['A1'].fill = fill(C_BLUE)
            ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[1].height = 24

            # 联赛信息
            ws.append(['LS联赛ID:', lr['ls_tournament_id'], 'LS联赛名:', lr['ls_name'],
                       'LS地区:', lr['ls_category'], '',
                       'TS联赛ID:', lr['ts_competition_id'], 'TS联赛名:', lr['ts_name'],
                       'TS国家:', lr['ts_country'], '相似分:', lr['league_score']])
            for cell in ws[2]:
                cell.fill = fill('D9E1F2')
                cell.font = Font(bold=True)

            ws.append(['LS 2026比赛数:', lr['ls_event_count_2026'], '', '', '', '', '',
                       'TS 2026比赛数:', lr.get('ts_event_count_2026', 0), '', '', '', '',
                       '已匹配:', lr.get('event_matched_count', 0)])
            ws.append([])

            # 表头（左LS 右TS）
            ws.append([
                'LS赛事ID', 'LS赛程时间', 'LS主队ID', 'LS主队名称', 'LS客队ID', 'LS客队名称',
                '队名相似分', '匹配级别', '已匹配',
                'TS赛事ID', 'TS赛程时间', 'TS主队ID', 'TS主队名称', 'TS客队ID', 'TS客队名称',
            ])
            hrow = ws.max_row
            for cell in ws[hrow]:
                cell.fill = fill(C_LIGHT)
                cell.font = hdr_font()
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.row_dimensions[hrow].height = 28

            # 比赛数据
            for em in ems:
                ws.append([
                    em['ls_event_id'], em['ls_scheduled'],
                    em['ls_home_id'], em['ls_home_name'],
                    em['ls_away_id'], em['ls_away_name'],
                    em['team_score'], em['match_level'], 'YES' if em['matched'] else 'NO',
                    em['ts_match_id'], em['ts_match_time_str'],
                    em['ts_home_id'], em['ts_home_name'],
                    em['ts_away_id'], em['ts_away_name'],
                ])
                ef = ev_fill.get(em['match_level'], fill(C_NO))
                for cell in ws[ws.max_row]:
                    cell.fill = ef
                    cell.alignment = Alignment(horizontal='center')
                for ci in [4, 6, 13, 15]:
                    ws.cell(ws.max_row, ci).alignment = Alignment(horizontal='left')

            auto_width(ws)

    wb.save(path)
    ts(f"Excel已保存: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    output = '/home/ubuntu/lsports_ts_match_2026.xlsx'
    ts("开始2026年LSports→TheSports匹配任务 (v3)")

    fb = run_sport_match('football')
    bb = run_sport_match('basketball')

    ts("导出Excel报表...")
    export_excel(fb, bb, output)

    ts(f"\n任务完成！文件: {output}")
